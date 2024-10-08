from flask import Flask, render_template, request, redirect, url_for, flash, Response, send_file
import os, platform, subprocess
import openpyxl
from openpyxl import load_workbook
import requests
from fpdf import FPDF
from ics import Calendar, Event
#from datetime import datetime, timedelta
import datetime
import sys
import re
from openpyxl import load_workbook
import os
import uuid
from flask import jsonify
import locale
import pytz

app = Flask(__name__)
app.secret_key = "secret_key"  # Für Flash-Messages

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xls', 'xlsx'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/')
def index():
    return render_template('index.html')

class ExcelProcessor:

    # Liste der Feiertage als Klassenvariable
    feiertage = [
        datetime.datetime(2023, 1, 1), datetime.datetime(2023, 4, 7), datetime.datetime(2023, 4, 10),
        datetime.datetime(2023, 5, 1), datetime.datetime(2023, 5, 18), datetime.datetime(2023, 5, 29),
        datetime.datetime(2023, 10, 3), datetime.datetime(2023, 10, 31), datetime.datetime(2023, 12, 25),
        datetime.datetime(2023, 12, 26), datetime.datetime(2024, 1, 1), datetime.datetime(2024, 3, 29),
        datetime.datetime(2024, 4, 1), datetime.datetime(2024, 5, 1), datetime.datetime(2024, 5, 9),
        datetime.datetime(2024, 5, 20), datetime.datetime(2024, 10, 3), datetime.datetime(2024, 10, 31),
        datetime.datetime(2024, 12, 25), datetime.datetime(2024, 12, 26), datetime.datetime(2025, 1, 1),
        datetime.datetime(2025, 4, 18), datetime.datetime(2025, 4, 21), datetime.datetime(2025, 5, 1),
        datetime.datetime(2025, 5, 29), datetime.datetime(2025, 6, 9), datetime.datetime(2025, 10, 3),
        datetime.datetime(2025, 10, 31), datetime.datetime(2025, 12, 25), datetime.datetime(2025, 12, 26)
    ]

    def __init__(self, filepath):
        self.filename = filepath 
        self.workbook = load_workbook(filepath)

        # Definiere relevante Tabellenblätter
        self.relevant_sheets = [sheet for sheet in self.workbook.sheetnames if
                                re.match(r'^KW\d{1,2}$', sheet)]  # Hinzugefügt

        #  Erstelle das Verzeichnis "tempfiles", falls es nicht existiert.
        if not os.path.exists('tempfiles'):
            os.makedirs('tempfiles')

        self.unique_id = uuid.uuid4().hex
        self.temp_pdf = f"tempfiles/temp_schedule_{self.unique_id}.pdf"
        self.modified_file = f"tempfiles/modified_file_{self.unique_id}.xlsx"

        # Aufrufen der cleanup Funktion bei der Initialisierung
        self.cleanup_old_files()

    def cleanup_old_files(self):
        # Liste der Ordner, in denen Dateien gelöscht werden sollen
        folders = ["tempfiles", "uploads"]

        # Aktuelles Datum und Zeitpunkt von vor 24 Stunden bestimmen
        now = datetime.datetime.now()
        one_day_ago = now - datetime.timedelta(days=1)

        for folder in folders:
            # Überprüfen, ob der Ordner existiert
            if os.path.exists(folder):
                print(f"Ordner {folder} wurde gefunden.")
                for filename in os.listdir(folder):
                    filepath = os.path.join(folder, filename)
                    # Nur Dateien überprüfen (keine Unterverzeichnisse)
                    if os.path.isfile(filepath):
                        file_creation_time = datetime.datetime.fromtimestamp(os.path.getctime(filepath))
                        print(f"Datei gefunden: {filepath}. Erstellt am: {file_creation_time}")

                        if file_creation_time < one_day_ago:
                            print(f"Die Datei {filepath} ist älter als 24 Stunden und wird gelöscht.")
                            os.remove(filepath)
                        else:
                            print(f"Die Datei {filepath} ist nicht älter als 24 Stunden.")
            else:
                print(f"Ordner {folder} wurde nicht gefunden!")

    def start_analysis(self):
        employees = set()
        if self.filename:
            self.workbook = openpyxl.load_workbook(self.modified_file)
            for sheet_name in self.relevant_sheets:
                sheet = self.workbook[sheet_name]
                letter_combinations = [sheet.cell(row=row, column=1).value for row in range(76, 138) if
                                       sheet.cell(row=row, column=1).value]
                for row in sheet.iter_rows(min_row=3, max_row=73, min_col=4, max_col=10):
                    for cell in row:
                        if cell.value:
                            for value in cell.value.split('/'):
                                for letter_combination in letter_combinations:
                                    if letter_combination[1:-1] in value:
                                        employees.add(letter_combination[1:-1])
        return sorted(employees)

    def replace_references_with_values(self):
        all_sheets = self.workbook.sheetnames
        for sheet_name in all_sheets:
            sheet = self.workbook[sheet_name]
            for row in sheet.iter_rows(min_row=3, max_row=73, min_col=4, max_col=10):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and re.match(r'^=.*$', cell.value):
                        # Prüfung auf ungültige Referenz (#REF!)
                        if '#REF!' in cell.value:
                            print(f"Ungültige Referenz gefunden in Zelle {cell.coordinate} im Blatt '{sheet_name}'")
                            continue  # Überspringen der aktuellen Zelle

                        reference = cell.value[1:]
                        if '!' in reference:
                            reference_sheet_name, reference_cell = reference.split('!', 1)
                            reference_sheet_name = reference_sheet_name.strip("'")
                        else:
                            reference_sheet_name = sheet_name
                            reference_cell = reference
                        try:
                            reference_sheet = self.workbook[reference_sheet_name]
                            reference_value = reference_sheet[reference_cell].value
                            cell.value = reference_value
                        except KeyError:
                            print(f"Worksheet '{reference_sheet_name}' not found. Available sheets: {all_sheets}")
                        except ValueError as e:
                            print(
                                f"Fehler beim Verarbeiten der Referenz in Zelle {cell.coordinate} im Blatt '{sheet_name}': {e}")
        self.workbook.save(self.modified_file)

    def check_plausibility(self):
        # Definiere die erwarteten Inhalte für jede Zelle.
        expected_contents = {
            'A15': 'Eingriff',
            'A24': 'POB',
            'A30': 'Inten',
            'A41': 'BD 1',
            'A42': 'BD 2',
        }

        # Iteriere über die relevanten Tabellenblätter.
        for sheet_name in self.relevant_sheets:
            # Hole das Tabellenblatt aus der Arbeitsmappe.
            sheet = self.workbook[sheet_name]

            # Iteriere über die erwarteten Inhalte.
            for cell, expected_content in expected_contents.items():
                # Hole den tatsächlichen Inhalt der Zelle.
                actual_content = sheet[cell].value

                # Wenn der tatsächliche Inhalt nicht den erwarteten Inhalt enthält, zeige eine Fehlermeldung an und gebe False zurück.
                if not actual_content or expected_content not in actual_content:
                    print("Fehler: Offenbar wurde die Sortierung der Wochenpläne verändert oder du hast einen alten Dienstlan geladen. Eine verlässliche Extraktion der Dienste ist nicht gewährleistet. Bitte informiere den Entwickler")
                    return False

        # Wenn alle Überprüfungen bestanden wurden, gebe True zurück.
        return True

    def show_schedule(self, selected_employee):
        # Setze die Lokalisierung auf Deutsch, falls verfügbar.
        def set_locale(category, loc):
            try:
                locale.setlocale(category, loc)
            except locale.Error:
                print(f"Locale {loc} not supported. Using default locale.")

        set_locale(locale.LC_TIME, 'de_DE.UTF-8')

        # Hole die abgedeckten Tage aus der Arbeitsmappe.
        covered_days = get_covered_days(self.workbook, self.relevant_sheets)

        # Hole den Monat und das Jahr vom zehnten abgedeckten Tag.
        tenth_day = datetime.datetime.strptime(covered_days[9].split()[0], '%d.%m.%Y')
        month_year = tenth_day.strftime('%B %Y')
        month = tenth_day.month
        year = tenth_day.year

        # Initialisiere ein Wörterbuch, um den Zeitplan des ausgewählten Mitarbeiters zu speichern.
        employee_schedule = {}

        # Iteriere über die relevanten Tabellenblätter.
        for sheet_name in self.relevant_sheets:
            # Hole das Tabellenblatt aus der Arbeitsmappe.
            sheet = self.workbook[sheet_name]

            # Iteriere über die Spalten des Tabellenblatts.
            for col in range(4, 11):
                # Hole das Datum aus der zweiten Zeile des Tabellenblatts.
                date = sheet.cell(row=2, column=col).value

                # Prüfe, ob das Datum abgedeckt ist und zum Monat und Jahr gehört.
                if date and date.strftime('%d.%m.%Y %A') in covered_days:
                    if date.month == month and date.year == year:
                        # Das Datum wird für den Zeitplan verwendet.
                        service = self.get_service(sheet, selected_employee, col)
                        if date in self.feiertage:
                            service += ' (Feiertag)'
                        employee_schedule[date.strftime('%d.%m.%Y')] = service
                        print(f"Tag verwendet: {date.strftime('%d.%m.%Y')}")
                    else:
                        # Das Datum wird nicht für den Zeitplan verwendet.
                        print(f"Tag ignoriert (nicht im Zielmonat/-jahr): {date.strftime('%d.%m.%Y')}")

        # Initialisiere einen String, um den Zeitplan des ausgewählten Mitarbeiters zu speichern.
        schedule_text = f"Dienstplan für {selected_employee} - {month_year}\n\n"

        # Definiere eine Liste von Wochentagen auf Deutsch.
        weekdays_german = ['Montag', 'Dienstag', 'Mittwoch', 'Donnerstag', 'Freitag', 'Samstag', 'Sonntag']

        # Hänge den formatierten Mitarbeiterzeitplan an den schedule_text-String an.
        for i, (date, service) in enumerate(employee_schedule.items()):
            # Hole den Wochentag auf Deutsch.
            weekday = weekdays_german[datetime.datetime.strptime(date, '%d.%m.%Y').weekday()]
            line = f"{date} ({weekday}): {service}\n"
            schedule_text += line

            # Füge eine Zeile nach jedem Sonntag ein.
            if weekday == 'Sonntag':
                schedule_text += '-' * 50 + '\n'

        # Füge den Haftungsausschluss-Text ein.
        disclaimer = "\nBitte überprüft den Plan auf seine Richtigkeit. Achtet im offiziellen Plan auf kurzfristige Änderungen!\n"
        schedule_text += disclaimer

        return schedule_text


    def get_service(self, sheet, employee, col):
        # Definiere ein Wörterbuch, um die Zeilennummern den Diensten zuzuordnen.
        services = {
            3: 'OP-Koordination',
            4: 'FD-OP',
            5: 'FD-OP',
            6: 'FD-OP',
            7: 'FD-OP',
            8: 'FD-OP',
            9: 'FD-OP',
            10: 'FD-OP',
            11: 'FD-OP',
            12: 'FD-lang',
            13: '?',
            14: 'FD-Außenbezirke',
            15: 'EGR-OA',
            16: 'FD-EGR',
            17: 'FD-EGR',
            18: 'FD-EGR',
            19: 'FD-BronchoHKL',
            20: 'FD-Geb',
            21: 'SD930',  # Direkte Zuweisung von 'SD930' zu Zeile 21
            22: 'SD11',
            23: 'SD13',
            24: 'POBE',
            25: 'Prämed',
            26: 'Prämed',
            27: 'Prämed',
            28: 'OA-ZOP',
            29: 'FD-Einarbeitung',
            30: 'FD-OA-Int',
            31: 'FD-FA-Int',
            32: 'FD-Int',
            33: 'FD-Int',
            34: 'SD-Int',
            35: 'SD-Int',
            36: 'ND-Int',
            37: 'NEF-Tag',
            38: 'NEF-Nacht',
            39: 'ITW-Tag',
            40: 'ITW-Nacht',
            41: 'BD1',
            42: 'BD2',
            43: 'Rufdienst',
        }

        # Füge den Dienst 'Frei' für die Zeilen 44 bis 73 hinzu.
        for row in range(44, 74):
            services[row] = 'Frei'

        # Spezielle Logik für Zeile 73
        def check_special_case(cell_value, current_date):
            if "unterr" in cell_value.lower():
                if current_date.weekday() == 0:  # Montag
                    return "SAN-Unterricht"
                elif current_date.weekday() == 2:  # Mittwoch
                    return "PJ-Unterricht"
            return "Frei"

        # Hole das Datum aus der aktuellen Spalte und bestimme den Dienst für Zeile 12.
        current_date = sheet.cell(row=2, column=col).value

        # Überprüfe, ob das aktuelle Datum ein Freitag ist.
        if current_date and current_date.weekday() == 4:  # Wenn es Freitag ist
            services[12] = 'SD10'
        else:
            services[12] = 'FD-lang'

        # Initialisiere eine Liste, um die Dienste des Mitarbeiters zu speichern.
        employee_services = []

        # Erkenne Tag/Nacht Dienst anhand der Position relativ zum /-Zeichen.
        for row in range(3, 74):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value:
                for value in cell_value.split('/'):
                    if employee in value:
                        if row == 73:
                            service = check_special_case(cell_value, current_date)
                        else:
                            service = services[row]
                        if row in [40, 41] and '/' in cell_value:
                            if cell_value.index(employee) == 0:
                                service += '/Tag'
                            else:
                                service += '/Nacht'
                        employee_services.append(service)

        # Wenn der Mitarbeitername in keiner Zeile gefunden wird.
        if not employee_services:
            date = sheet.cell(row=2, column=col).value

            # Überprüfen, ob das Datum ein Samstag, Sonntag oder ein Feiertag ist
            if date and (date.weekday() in [5, 6] or date in self.feiertage):
                employee_services.append('Frei')
            else:
                employee_services.append('?')

        return ', '.join(employee_services)


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        flash('Keine Datei ausgewählt')
        return redirect(request.url)
    file = request.files['file']
    if file.filename == '':
        flash('Keine Datei ausgewählt')
        return redirect(request.url)
    if file and allowed_file(file.filename):
        filename = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(filename)

        # Lade die Arbeitsmappe.
        workbook = load_workbook(filename)

        # Definiere die relevanten Tabellenblätter basierend auf dem KW-Muster.
        relevant_sheets = [sheet for sheet in workbook.sheetnames if re.match(r'^KW\d{1,2}$', sheet)]

        # Erstelle eine Instanz von ExcelProcessor und ersetze die Verweise.
        processor = ExcelProcessor(filename)
        processor.relevant_sheets = relevant_sheets
        processor.replace_references_with_values()

        # Starte die Analyse, um die Mitarbeiter zu erhalten.
        employees = processor.start_analysis()

        # Gib die UID und die Mitarbeiterliste als JSON zurück:
        return jsonify(uid=processor.unique_id, employees=employees)

    else:
        flash('Nicht unterstütztes Dateiformat')
        return redirect(request.url)



def get_covered_days(workbook, relevant_sheets):
    #  Initialisiere eine Liste, um die abgedeckten Tage zu speichern.
    covered_days = []

    # Iteriere über die relevanten Tabellenblätter.
    for sheet_name in relevant_sheets:
        # Hole das Tabellenblatt aus der Arbeitsmappe.
        sheet = workbook[sheet_name]

        # Hole die Daten aus der zweiten Zeile des Tabellenblatts.
        dates = [sheet.cell(row=2, column=col).value for col in range(4, 11)]

        # Füge die Daten der abgedeckten Tage der covered_days-Liste hinzu.
        covered_days.extend(dates)

        # Konvertiere die Daten in das gewünschte Format.
    covered_days = [date.strftime('%d.%m.%Y %A') for date in covered_days if isinstance(date, datetime.datetime)]

    return covered_days


@app.route('/handle_employee_selection', methods=['POST'])
def handle_employee_selection():
    data = request.json
    selected_employee = data.get('employee')
    uid = data.get('uid')

    # Drucken der ausgewählten Mitarbeiter in die Konsole
    print(f"Ausgewählter Mitarbeiter: {selected_employee}")

    # Erstellet den Prozessor mit der Datei, die der UID entspricht:
    filepath = f"tempfiles/modified_file_{uid}.xlsx"
    processor = ExcelProcessor(filepath)

    # Überprüfe die Plausibilität
    if not processor.check_plausibility():
        error_message = ("Fehler: Offenbar wurde die Sortierung der Wochenpläne verändert oder "
                         "du hast einen alten Dienstlan geladen. Eine verlässliche Extraktion der Dienste ist "
                         "nicht gewährleistet. Bitte informiere den Entwickler")
        return jsonify(message="Fehler bei der Plausibilitätsprüfung", schedule=error_message)

    #  Hole den Zeitplan für den ausgewählten Mitarbeiter.
    schedule = processor.show_schedule(selected_employee)

    # Gib den Zeitplan in der Antwort zurück.
    return jsonify(message="Erfolgreich verarbeitet", schedule=schedule)


@app.route('/generate_pdf', methods=['POST'])
def generate_pdf():
    # Definition der Schichtzeiten
    shift_times = {
        "OP-Koordination": {"start": "07:20", "end": "15:50"},
        "FD-OP": {"start": "07:20", "end": "15:50"},
        "FD-lang": {"start": "07:20", "end": "18:05"},
        "FD-Außenbezirke": {"start": "07:20", "end": "15:50"},
        "EGR-OA": {"start": "07:20", "end": "15:50"},
        "FD-EGR": {"start": "07:20", "end": "15:50"},
        "FD-BronchoHKL": {"start": "07:20", "end": "15:50"},
        "FD-Geb": {"start": "07:20", "end": "15:50"},
        "SD930": {"start": "09:30", "end": "18:00"},
        "SD11": {"start": "11:00", "end": "19:30"},
        "SD13": {"start": "13:00", "end": "21:30"},
        "POBE": {"start": "10:00", "end": "18:30"},
        "Prämed": {"start": "07:20", "end": "15:50"},
        "OA-ZOP": {"start": "07:20", "end": "15:50"},
        "FD-Einarbeitung": {"start": "07:20", "end": "15:50"},
        "FD-OA-Int": {"start": "06:50", "end": "15:20"},
        "FD-FA-Int": {"start": "06:50", "end": "15:20"},
        "FD-FA-Int, FD-Int": {"start": "06:50", "end": "15:20"},
        "FD-Int": {"start": "06:50", "end": "15:20"},
        "SD-Int": {"start": "14:20", "end": "22:50"},
        "ND-Int": {"start": "21:50", "end": "07:50"},  # Über Mitternacht
        "NEF-Tag": {"start": "06:50", "end": "19:35"},
        "NEF-Nacht": {"start": "18:50", "end": "07:35"},  # Über Mitternacht
        "BD1": {"start": "15:40", "end": "07:40"},  # Über Mitternacht
        "BD2": {"start": "15:40", "end": "07:40"},  # Über Mitternacht
        "SD10": {"start": "10:00", "end": "18:30"},  
        "BD1/Tag": {"start": "08:00", "end": "20:00"},
        "BD1/Nacht": {"start": "20:00", "end": "08:00"},  # Über Mitternacht
        "BD2/Tag": {"start": "08:00", "end": "20:00"},
        "BD2/Nacht": {"start": "20:00", "end": "08:00"}  # Über Mitternacht
    }
  
    weekend_shift_times = {
        "FD-Int": {"start": "06:50", "end": "17:20"},
        "SD-Int": {"start": "16:20", "end": "22:50"}
    }


    def get_shift_time(service, date_str):
      """
      Ermittle die Schichtzeiten für einen gegebenen Dienst und Datum.
      """
      day_of_week = datetime.datetime.strptime(date_str, "%d.%m.%Y").weekday()
  
      # Wochentag oder Wochenende?
      shift_time_dict = weekend_shift_times if day_of_week >= 5 and service in weekend_shift_times else shift_times
  
      return shift_time_dict.get(service, {"start": "(ganztägig)", "end": ""})

    # PDF-Erstellung
    schedule_text = request.json.get("schedule_text", "")
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=8)

    # Deaktiviere automatischen Seitenumbruch
    pdf.set_auto_page_break(auto=False)
    
    lines = schedule_text.split('\n')
    for line in lines:
        if 'Samstag' in line or 'Sonntag' in line:
            pdf.set_fill_color(200, 200, 200)  # Hintergrund für Wochenendtage
        else:
            pdf.set_fill_color(255, 255, 255)  # Hintergrund für Wochentage
  
        parts = line.split(": ")
        if len(parts) == 2:
            date_str, service = parts
            date_str_for_shift_time = date_str.split(' ')[0].strip()
            shift_time = get_shift_time(service, date_str_for_shift_time)
            if shift_time['start'] == "(ganztägig)":
                time_info = " (ganztägig)"
            else:
                time_info = f" ({shift_time['start']} - {shift_time['end']})" if shift_time['start'] else ""

  
            # Dienst in schwarzer Schrift
            pdf.set_text_color(0, 0, 0)  # Schwarze Farbe für den Dienst
            pdf.cell(35, 5, date_str + ":", ln=False, fill=True)  # Breite für Datum
            pdf.cell(60, 5, service, ln=False, fill=True)  # Breite für Dienst
  
            # Uhrzeiten in grauer Schrift
            pdf.set_text_color(100, 100, 100)  # Graue Farbe für die Uhrzeiten
            pdf.cell(40, 5, time_info, ln=True, fill=True)  # Breite für Uhrzeiten
        else:
            # Für Zeilen ohne Dienst
            pdf.set_text_color(0, 0, 0)  # Schwarze Schriftfarbe
            pdf.multi_cell(0, 5, line, fill=True)


    # Hinzufügen des Erstellungsdatums
    adjusted_time = datetime.datetime.now() + datetime.timedelta(hours=2)
    creation_date = adjusted_time.strftime('%d.%m.%Y')
    creation_time = adjusted_time.strftime('%H:%M')
    pdf.multi_cell(0, 10, f"Erstellt am {creation_date} um {creation_time} Uhr", align='L')
  
    # Fußzeilentext definieren
    footer_text = "- DeinDienstplan ist ein Webservice und verursacht Betriebskosten. Wenn das Programm dir nützlich ist, erwäge bitte eine Spende. -"
  
    # Fußzeilentext hinzufügen
    footer_y_position = 297 - 20  # 297 mm ist die Höhe einer A4-Seite, 20 mm Abstand vom unteren Rand
    pdf.set_y(footer_y_position)
    pdf.set_font("Arial", size=6)
    pdf.cell(0, 10, footer_text, 0, 0, 'C')

    # Log-Aktion
    firstLine = schedule_text.splitlines()[0]
    log_action(firstLine, "PDF erstellt")

    # PDF-Inhalt als Antwort senden
    pdf_content = pdf.output(dest='S').encode('latin1')
    response = Response(pdf_content, content_type='application/pdf')
    return response

@app.route('/generate_ics', methods=['POST'])
def generate_ics():
    data = request.json
    schedule_text = data['schedule_text']
    event_type = data.get('event_type', 'fullDay')  # Standardmäßig 'fullDay'

    c = Calendar()

    timezone = pytz.timezone("Europe/Berlin")

    shift_times = {
        "OP-Koordination": {"start": "07:20", "end": "15:50"},
        "FD-OP": {"start": "07:20", "end": "15:50"},
        "FD-lang": {"start": "07:20", "end": "18:05"},
        "FD-Außenbezirke": {"start": "07:20", "end": "15:50"},
        "EGR-OA": {"start": "07:20", "end": "15:50"},
        "FD-EGR": {"start": "07:20", "end": "15:50"},
        "FD-BronchoHKL": {"start": "07:20", "end": "15:50"},
        "FD-Geb": {"start": "07:20", "end": "15:50"},
        "SD930": {"start": "09:30", "end": "18:00"},
        "SD11": {"start": "11:00", "end": "19:30"},
        "SD13": {"start": "13:00", "end": "21:30"},
        "POBE": {"start": "10:00", "end": "18:30"},
        "Prämed": {"start": "07:20", "end": "15:50"},
        "OA-ZOP": {"start": "07:20", "end": "15:50"},
        "FD-Einarbeitung": {"start": "07:20", "end": "15:50"},
        "FD-OA-Int": {"start": "06:50", "end": "15:20"},
        "FD-FA-Int, FD-Int": {"start": "06:50", "end": "15:20"},
        "FD-FA-Int": {"start": "06:50", "end": "15:20"},
        "FD-Int": {"start": "06:50", "end": "15:20"},
        "SD-Int": {"start": "14:20", "end": "22:50"},
        "ND-Int": {"start": "21:50", "end": "07:50"},
        "NEF-Tag": {"start": "06:50", "end": "19:35"},
        "NEF-Nacht": {"start": "18:50", "end": "07:35"},
        "ITW-Tag": {"start": "06:50", "end": "19:35"},
        "ITW-Nacht": {"start": "18:50", "end": "07:35"},
        "BD1": {"start": "15:40", "end": "07:40"},
        "BD2": {"start": "15:40", "end": "07:40"},
        "SD10": {"start": "10:00", "end": "18:30"},
        "BD1/Tag": {"start": "08:00", "end": "20:00"},
        "BD1/Nacht": {"start": "20:00", "end": "08:00"},
        "BD2/Tag": {"start": "08:00", "end": "20:00"},
        "BD2/Nacht": {"start": "20:00", "end": "08:00"}
    }
  
    weekend_shift_times = {
        "FD-Int": {"start": "06:50", "end": "17:20"},
        "SD-Int": {"start": "16:20", "end": "22:50"}
    }


    lines = schedule_text.split('\n')
    for line in lines:
      parts = line.split(": ")
      if len(parts) == 2:
          date_str, service = parts
          date_parts = date_str.split(" ")
          if len(date_parts) == 2:
              date, weekday = date_parts
              day, month, year = date.split('.')
              formatted_date = f"{year}-{month}-{day}"
              e = Event()
              e.name = service
              if event_type == 'shiftTimes' and service in shift_times:
                  day_of_week = datetime.datetime.strptime(formatted_date, "%Y-%m-%d").weekday()

                  if day_of_week >= 5 and service in weekend_shift_times:
                      start_time = weekend_shift_times[service]["start"]
                      end_time = weekend_shift_times[service]["end"]
                  else:
                      start_time = shift_times[service]["start"]
                      end_time = shift_times[service]["end"]

                  start_datetime = datetime.datetime.strptime(f"{formatted_date} {start_time}", "%Y-%m-%d %H:%M")
                  end_datetime = datetime.datetime.strptime(f"{formatted_date} {end_time}", "%Y-%m-%d %H:%M")

                  # Apply timezone
                  start_datetime = timezone.localize(start_datetime)
                  end_datetime = timezone.localize(end_datetime)

                  # Adjust for over-midnight shifts
                  if end_datetime <= start_datetime:
                      end_datetime += datetime.timedelta(days=1)

                  e.begin = start_datetime
                  e.end = end_datetime
              else:
                  # For all-day events, just use the date without timezone conversion
                  all_day_date = datetime.datetime.strptime(formatted_date, "%Y-%m-%d")
                  e.begin = all_day_date
                  e.make_all_day()  # Mark as an all-day event
              c.events.add(e)

    employee_name = lines[0].split(' ')[2] if len(lines) > 0 else "Unbekannt"
    month_year = ' '.join(lines[0].split(' ')[-2:]) if len(lines) > 0 else "Unbekannt"
    ics_filename = f"Dienstplan-{employee_name}-{month_year}.ics"

    ics_content = c.serialize()
    with open(ics_filename, 'w', encoding='utf-8') as my_file:
        my_file.write(ics_content)

    firstLine = schedule_text.splitlines()[0]
    log_action(firstLine, "ICS erstellt")

    return send_file(ics_filename, as_attachment=True, download_name=ics_filename)


def log_action(schedule_first_line, filetype):
    # IFTTT-Webhook-URL
    webhook_url = 'https://maker.ifttt.com/trigger/Dienstplan/with/key/jeYCdecKNOCcFc5NRUWu07ufuHSQKwrnyCuA_eQZZN7'

    # Daten für den Webhook
    data = {
        "value1": schedule_first_line,
        "value2": filetype
    }

    # Senden des Webhooks
    response = requests.post(webhook_url, json=data) 
    return response

if __name__ == '__main__':
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
    app.run(host='0.0.0.0', debug=False)

